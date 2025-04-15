from rest_framework.views import APIView
from rest_framework.response import Response
from django.http import HttpResponse
import openpyxl
from openpyxl.utils import get_column_letter
from io import BytesIO
from ..models import Participant
from ..utils import generate_title_number, generate_qrcode


class HabilElect:
    objects = None


class ExportHabilitationExcelView(APIView):
    """
    Vue DRF pour exporter les habilitations en fichier Excel.
    """
    authentication_classes = []
    permission_classes = []

    def get(self, request, *args, **kwargs):
        # Créer un nouveau fichier Excel
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Habilitations"

        # En-têtes de colonnes
        headers = [
            'Nom et Prénom',
            'Fonction',
            'Adresse Entreprise',
            'Adresse Entreprise 2',
            'Durée de Formation',
            'Date',
            'Numéro de Titre',
            'Symboles',
            'Symboles 2',
            'Domaine de tension',
            'Personnel',
            'Signature',
            'Titre',
            'Lieu',
            'Nom',
            'Prénom',
            'Fonction 2',
            'Nom Employeur',
            'Fonction Employeur',
            'Prénom Employeur',
            'Référence Employeur',
            'Date Employeur',
            'Validité',
            'Référence',
            'QR Code',
            'QR Code 2',
            'Photo',
            'Installations Concernées',
            'Indications',
            'Logo'
        ]
        ws.append(headers)

        # Récupérer les habilitations avec leurs relations
        habilitations = HabilElect.objects.prefetch_related(
            'participants', 'type_habilitation', 'tension', 'roles'
        )

        # Insérer les données dans les lignes du fichier Excel
        for habil in habilitations:
            for participant in habil.participants.all():
                types_habilitation = ", ".join([h.code for h in habil.type_habilitation.all()])
                tensions = ",\n".join([t.nom for t in habil.tension.all()])
                roles = ",\n".join([r.name for r in habil.roles.all()])
                Symboles2 = ",\n".join([t.code for t in habil.type_habilitation.all()])

                # Construire la ligne
                qrc = generate_qrcode(habil.titre)
                row = [
                    f"{participant.nom} {participant.prenom}",  # Nom et Prénom
                    participant.fonction,                      # Fonction
                    participant.employeur.name if participant.employeur else "",  # Adresse Entreprise
                    participant.employeur.name if participant.employeur else "",  # Adresse Entreprise 2
                    habil.duree,  # Durée de Formation
                    habil.date_debut.strftime("%d-%m-%y"),  # Date
                    generate_title_number(site=participant.employeur.id, index=participant.id), # Numéro de Titre
                    types_habilitation,                      # Symboles
                    Symboles2,                                # Symboles 2
                    tensions,                                # Domaine de tension
                    roles,                                   # Personnel
                    habil.responsable,                       # Signature
                    habil.titre,                       # Titre
                    habil.site,                              # Lieu
                    participant.nom,                         # Nom
                    participant.prenom,                      # Prénom
                    participant.fonction,                    # Fonction 2
                    participant.employeur.nom_responsable if participant.employeur else "",  # Nom Employeur
                    participant.employeur.fonction_emp,                # Fonction Employeur
                    participant.employeur.prenom_resp,                      # Prénom Employeur
                    participant.employeur.id,                # Référence Employeur
                    habil.date_creation.strftime("%d-%m-%y"),  # Date Employeur
                    habil.date_expiration.strftime("%d-%m-%y"),  # Validité
                    habil.id,                                # Référence
                    qrc,                           # QR Code
                    qrc,                         # QR Code 2
                    participant.photo.url if participant.photo else "",  # Photo
                    "Installations Concernées",              # Installations Concernées
                    "Indications",                           # Indications
                    participant.employeur.logo if participant.employeur.logo else ""                         # Logo
                ]
                ws.append(row)

        # Ajuster automatiquement la largeur des colonnes
        for col in ws.columns:
            max_length = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except Exception:
                    pass
            ws.column_dimensions[col_letter].width = max_length + 2

        # Enregistrer dans un fichier en mémoire
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        # Créer une réponse HTTP avec le fichier Excel
        response = HttpResponse(
            output,
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="habilitations.xlsx"'
        return response
